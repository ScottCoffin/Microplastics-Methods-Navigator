"""
Fill selected Crosswalk Table metadata from Zotero PDFs.

This script is dry-run by default. It reads Zotero locally, uses Zotero
abstractNote first, converts matched PDF attachments to Markdown with
pymupdf4llm only when needed, and limits PDF conversion to the first 5 pages
and first 500 Markdown lines. It falls back to DOI/URL landing pages for rows
without Zotero abstracts/PDFs, asks Claude Haiku for structured metadata, and
can write Particle Size Range, Key Metrics / Output, Abstract, PDF Available in
Zotero?, and Reviewed back to a workbook.

Command reference:
    Install metadata dependencies:
        python -m pip install -r metadata_requirements.txt

    Beta test one DOI, dry-run:
        python fill_crosswalk_metadata.py --doi 10.1016/j.chemosphere.2022.134282

    Write one DOI to crosswalk.metadata_filled.xlsx:
        python fill_crosswalk_metadata.py --doi 10.1016/j.chemosphere.2022.134282 --write

    Process all rows into a populated copy:
        python fill_crosswalk_metadata.py --write

    Resume from crosswalk.metadata_filled.xlsx in batches of 3:
        python fill_crosswalk_metadata.py --resume --batch-size 3

    Resume at a specific Excel row and process the next 3 eligible rows:
        python fill_crosswalk_metadata.py --resume --start-row 42 --batch-size 3

    Resume and keep looping until no unreviewed rows remain:
        python fill_crosswalk_metadata.py --resume --batch-size 3 --batch-loop

    After adding missing PDFs to Zotero, rerun only unreviewed rows:
        python fill_crosswalk_metadata.py --resume --batch-size 3 --batch-loop

    Disable DOI/URL landing-page scraping for an offline Zotero/PDF-only run:
        python fill_crosswalk_metadata.py --resume --batch-size 3 --skip-url-scrape

    Start a fresh audit log by deleting metadata_cache/crosswalk_metadata_audit.csv
    before running. Otherwise, audit rows append across batches/runs.

    Overwrite existing metadata fields when needed:
        python fill_crosswalk_metadata.py --resume --overwrite

Legacy examples:
    Example beta test:
    python fill_crosswalk_metadata.py --doi 10.1016/j.chemosphere.2022.134282

    Write a populated copy:
    python fill_crosswalk_metadata.py --doi 10.1016/j.chemosphere.2022.134282 --write
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import sqlite3
import sys
import tempfile
import textwrap
import urllib.error
import urllib.request
from copy import copy
from dataclasses import dataclass
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_MODEL = "claude-haiku-4-5-20251001"
HEADER_ROW = 2
DATA_START_ROW = 3
TARGET_HEADERS = {
    "particle_size_range": "Particle Size Range",
    "key_metrics_output": "Key Metrics / Output",
    "abstract": "Abstract",
    "pdf_available_zotero": "PDF Available in Zotero?",
    "reviewed": "Reviewed",
}
PDF_MAX_PAGES = 5
PDF_MAX_MARKDOWN_LINES = 500


@dataclass
class ZoteroItem:
    item_id: int
    key: str
    title: str | None
    doi: str | None
    abstract: str | None
    date: str | None
    attachment_paths: list[Path]


def clean_text(value: Any) -> str:
    text = unescape(str(value or ""))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


class LandingPageParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.meta: dict[str, str] = {}
        self.json_ld: list[str] = []
        self.text_parts: list[str] = []
        self._skip_depth = 0
        self._capture_json_ld = False
        self._json_ld_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_dict = {name.lower(): value or "" for name, value in attrs}
        tag = tag.lower()
        if tag in {"script", "style", "noscript"}:
            if tag == "script" and "ld+json" in attrs_dict.get("type", "").lower():
                self._capture_json_ld = True
                self._json_ld_parts = []
            else:
                self._skip_depth += 1
            return
        if tag == "meta":
            key = (
                attrs_dict.get("name")
                or attrs_dict.get("property")
                or attrs_dict.get("itemprop")
            )
            content = attrs_dict.get("content")
            if key and content:
                self.meta[key.lower()] = clean_text(content)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag == "script" and self._capture_json_ld:
            joined = "".join(self._json_ld_parts).strip()
            if joined:
                self.json_ld.append(joined)
            self._capture_json_ld = False
            self._json_ld_parts = []
            return
        if tag in {"script", "style", "noscript"} and self._skip_depth:
            self._skip_depth -= 1

    def handle_data(self, data: str) -> None:
        if self._capture_json_ld:
            self._json_ld_parts.append(data)
            return
        if self._skip_depth:
            return
        text = clean_text(data)
        if text:
            self.text_parts.append(text)


def normalize_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip().lower()


def normalize_doi(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("https://doi.org/", "").replace("http://doi.org/", "")
    text = text.replace("doi:", "").strip()
    match = re.search(r"10\.\d{4,9}/[^\s,;]+", text, flags=re.I)
    if not match:
        return None
    return match.group(0).rstrip(".").lower()


def normalize_title(value: Any) -> str:
    text = re.sub(r"[^a-z0-9]+", " ", str(value or "").lower())
    return " ".join(text.split())


def source_url_from_row(row: dict[str, Any]) -> str | None:
    raw = str(row.get("DOI/URL") or "").strip()
    if not raw:
        return None
    doi = normalize_doi(raw)
    if doi:
        return f"https://doi.org/{doi}"
    match = re.search(r"https?://\S+", raw)
    if not match:
        return None
    return match.group(0).rstrip(").,;")


def load_env_file(path: Path) -> None:
    """Load simple KEY=VALUE lines without adding a runtime dependency."""
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def workbook_headers(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(HEADER_ROW, col).value)
        if header:
            headers[header] = col
    return headers


def copy_column_style(ws, src_col: int, dst_col: int) -> None:
    for row in range(1, ws.max_row + 1):
        src = ws.cell(row, src_col)
        dst = ws.cell(row, dst_col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
    ws.column_dimensions[get_column_letter(dst_col)].width = max(
        ws.column_dimensions[get_column_letter(src_col)].width or 18,
        24,
    )


def ensure_target_columns(ws) -> dict[str, int]:
    headers = workbook_headers(ws)
    columns: dict[str, int] = {}

    for key, label in TARGET_HEADERS.items():
        normalized = normalize_header(label)
        if normalized in headers:
            columns[key] = headers[normalized]
            continue

        new_col = ws.max_column + 1
        copy_column_style(ws, max(1, new_col - 1), new_col)
        ws.cell(1, new_col).value = "Extended Metadata"
        ws.cell(HEADER_ROW, new_col).value = label
        columns[key] = new_col
        headers[normalized] = new_col

    return columns


def collection_ids(con: sqlite3.Connection, names: list[str], ids: list[int]) -> list[int]:
    found = set(ids)
    cur = con.cursor()
    for name in names:
        rows = cur.execute(
            "select collectionID from collections where collectionName = ?",
            (name,),
        ).fetchall()
        found.update(row[0] for row in rows)
    return sorted(found)


def zotero_items(con: sqlite3.Connection, storage_dir: Path, collection_ids_: list[int]) -> list[ZoteroItem]:
    if not collection_ids_:
        return []

    placeholders = ",".join("?" for _ in collection_ids_)
    cur = con.cursor()
    rows = cur.execute(
        f"""
        with coll_items as (
            select distinct itemID
            from collectionItems
            where collectionID in ({placeholders})
        ),
        item_field as (
            select d.itemID, f.fieldName, v.value
            from itemData d
            join fields f on f.fieldID = d.fieldID
            join itemDataValues v on v.valueID = d.valueID
            where f.fieldName in ('title', 'DOI', 'abstractNote', 'date')
        ),
        piv as (
            select itemID,
                   max(case when fieldName='title' then value end) title,
                   max(case when fieldName='DOI' then value end) doi,
                   max(case when fieldName='abstractNote' then value end) abstract,
                   max(case when fieldName='date' then value end) date
            from item_field
            group by itemID
        )
        select i.itemID, i.key, p.title, p.doi, p.abstract, p.date
        from coll_items ci
        join items i on i.itemID = ci.itemID
        left join piv p on p.itemID = i.itemID
        """,
        collection_ids_,
    ).fetchall()

    items: list[ZoteroItem] = []
    for item_id, key, title, doi, abstract, date in rows:
        attachments = attachment_paths(cur, item_id, storage_dir)
        items.append(
            ZoteroItem(
                item_id=item_id,
                key=key,
                title=title,
                doi=normalize_doi(doi),
                abstract=abstract,
                date=date,
                attachment_paths=attachments,
            )
        )
    return items


def attachment_paths(cur: sqlite3.Cursor, parent_item_id: int, storage_dir: Path) -> list[Path]:
    rows = cur.execute(
        """
        select child.key, a.path, a.contentType
        from itemAttachments a
        join items child on child.itemID = a.itemID
        where a.parentItemID = ?
        """,
        (parent_item_id,),
    ).fetchall()
    paths: list[Path] = []
    for key, path, content_type in rows:
        if content_type != "application/pdf" or not path:
            continue
        if path.startswith("storage:"):
            candidate = storage_dir / key / path.split(":", 1)[1]
        else:
            candidate = Path(path)
        if candidate.exists():
            paths.append(candidate)
    return paths


def choose_item(row: dict[str, Any], items: list[ZoteroItem]) -> ZoteroItem | None:
    row_doi = normalize_doi(row.get("DOI/URL"))
    if row_doi:
        doi_matches = [item for item in items if item.doi == row_doi]
        if doi_matches:
            return sorted(
                doi_matches,
                key=lambda item: (
                    not bool(item.attachment_paths),
                    not bool(item.abstract),
                    item.item_id,
                ),
            )[0]

    title = normalize_title(row.get("Full Title"))
    if title:
        title_matches = [item for item in items if normalize_title(item.title) == title]
        if title_matches:
            return sorted(
                title_matches,
                key=lambda item: (not bool(item.attachment_paths), item.item_id),
            )[0]
    return None


def first_markdown_lines(markdown: str, max_lines: int = PDF_MAX_MARKDOWN_LINES) -> str:
    return "\n".join(markdown.splitlines()[:max_lines])


def convert_pdf_to_markdown(pdf_path: Path, cache_dir: Path) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    out_path = cache_dir / (
        f"{pdf_path.stem}.first{PDF_MAX_PAGES}pages_"
        f"first{PDF_MAX_MARKDOWN_LINES}lines.md"
    )
    if out_path.exists() and out_path.stat().st_mtime >= pdf_path.stat().st_mtime:
        return out_path

    try:
        import fitz
        import pymupdf4llm
    except ImportError as exc:
        raise RuntimeError(
            "pymupdf4llm and PyMuPDF are required for PDF conversion. Install with: "
            "python -m pip install pymupdf4llm"
        ) from exc

    subset_path: Path | None = None
    source_path = pdf_path
    try:
        with fitz.open(pdf_path) as src:
            page_count = min(src.page_count, PDF_MAX_PAGES)
            if page_count <= 0:
                raise RuntimeError(f"PDF has no readable pages: {pdf_path}")
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                subset_path = Path(tmp.name)
            with fitz.open() as subset:
                subset.insert_pdf(src, from_page=0, to_page=page_count - 1)
                subset.save(subset_path)
            source_path = subset_path

        markdown = pymupdf4llm.to_markdown(str(source_path))
        out_path.write_text(first_markdown_lines(markdown), encoding="utf-8")
    finally:
        if subset_path and subset_path.exists():
            subset_path.unlink(missing_ok=True)
    return out_path


def jsonld_values(payload: Any, keys: set[str]) -> list[str]:
    values: list[str] = []
    if isinstance(payload, dict):
        for key, value in payload.items():
            if key.lower() in keys and isinstance(value, str):
                cleaned = clean_text(value)
                if cleaned:
                    values.append(cleaned)
            else:
                values.extend(jsonld_values(value, keys))
    elif isinstance(payload, list):
        for item in payload:
            values.extend(jsonld_values(item, keys))
    return values


def landing_page_cache_path(url: str, cache_dir: Path) -> Path:
    safe = re.sub(r"[^a-zA-Z0-9]+", "_", url).strip("_")[:140]
    return cache_dir / f"{safe or 'landing_page'}.txt"


def extract_relevant_visible_text(text_parts: list[str], max_chars: int) -> str:
    if not text_parts:
        return ""
    lines = [part for part in text_parts if len(part) > 2]
    joined = "\n".join(lines)
    section_pattern = re.compile(
        r"(?is)\b(abstract|scope|description|overview)\b\s*[:\n]\s*(.{80,2500})"
    )
    matches = []
    for match in section_pattern.finditer(joined):
        section = clean_text(match.group(0))
        if section:
            matches.append(section)
    if matches:
        return "\n\n".join(matches)[:max_chars]
    return "\n".join(lines[:80])[:max_chars]


def scrape_landing_page_context(
    url: str | None,
    cache_dir: Path,
    max_chars: int,
) -> tuple[str, str, str | None]:
    if not url:
        return "", "none", None
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = landing_page_cache_path(url, cache_dir)
    if cache_path.exists():
        return cache_path.read_text(encoding="utf-8", errors="replace")[:max_chars], "url_cache", str(cache_path)

    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (compatible; CrosswalkMetadataBot/1.0; "
                "+https://oehha.ca.gov/)"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        },
    )
    with urllib.request.urlopen(request, timeout=45) as response:
        content_type = response.headers.get("content-type", "")
        charset_match = re.search(r"charset=([\w.-]+)", content_type, flags=re.I)
        charset = charset_match.group(1) if charset_match else "utf-8"
        raw = response.read(2_000_000)
    html = raw.decode(charset, errors="replace")
    parser = LandingPageParser()
    parser.feed(html)

    candidates: list[str] = []
    for key in [
        "citation_abstract",
        "dc.description",
        "dcterms.description",
        "description",
        "og:description",
        "twitter:description",
    ]:
        value = parser.meta.get(key)
        if value:
            candidates.append(value)
    for block in parser.json_ld:
        try:
            candidates.extend(
                jsonld_values(
                    json.loads(block),
                    {"abstract", "description"},
                )
            )
        except json.JSONDecodeError:
            continue

    visible_context = extract_relevant_visible_text(parser.text_parts, max_chars)
    if visible_context:
        candidates.append(visible_context)

    deduped: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        cleaned = clean_text(candidate)
        key = cleaned.lower()
        if cleaned and key not in seen:
            seen.add(key)
            deduped.append(cleaned)
    context = "\n\n".join(deduped)[:max_chars]
    if context:
        cache_path.write_text(context, encoding="utf-8")
        return context, "url_landing_page", str(cache_path)
    return "", "url_no_context", None


def old_select_llm_context(markdown: str, max_chars: int) -> str:
    terms = [
        "abstract",
        "particle size",
        "size range",
        "micrometer",
        "µm",
        "um",
        "recovery",
        "accuracy",
        "precision",
        "interlaboratory",
        "metric",
        "output",
        "result",
    ]
    chunks = []
    lines = markdown.splitlines()
    window = 12
    for idx, line in enumerate(lines):
        low = line.lower()
        if any(term in low for term in terms):
            start = max(0, idx - window)
            end = min(len(lines), idx + window + 1)
            chunks.append("\n".join(lines[start:end]))

    front_matter = "\n".join(lines[:220])
    combined = front_matter + "\n\n--- MATCHED CONTEXT ---\n\n" + "\n\n---\n\n".join(chunks)
    return combined[:max_chars]


def extract_abstract_context(
    markdown: str,
    max_chars: int,
    fallback_lines: int = PDF_MAX_MARKDOWN_LINES,
) -> tuple[str, str]:
    lines = markdown.splitlines()
    abstract_start = None
    heading_pattern = re.compile(r"^\s{0,3}#{0,6}\s*(abstract|summary)\s*$", re.I)
    next_heading_pattern = re.compile(
        r"^\s{0,3}#{1,6}\s+\S+|^\s{0,3}(keywords?|introduction|background)\s*$",
        re.I,
    )

    for idx, line in enumerate(lines):
        if heading_pattern.match(line.strip()):
            abstract_start = idx
            break

    if abstract_start is not None:
        end = min(len(lines), abstract_start + fallback_lines)
        for idx in range(abstract_start + 1, end):
            if (
                next_heading_pattern.match(lines[idx].strip())
                and idx > abstract_start + 3
            ):
                end = idx
                break
        context = "\n".join(lines[abstract_start:end]).strip()
        if context:
            return context[:max_chars], "pdf_abstract_section"

    return "\n".join(lines[:fallback_lines])[:max_chars], "pdf_first_300_lines"


def resolve_llm_context(
    item: ZoteroItem | None,
    markdown: str | None,
    url_context: str | None,
    max_chars: int,
) -> tuple[str, str]:
    if item and item.abstract:
        return item.abstract[:max_chars], "zotero_abstractNote"
    if markdown:
        return extract_abstract_context(markdown, max_chars)
    if url_context:
        return url_context[:max_chars], "url_landing_page"
    return "", "none"


def build_prompt(
    row: dict[str, Any],
    zotero_item: ZoteroItem | None,
    context: str,
    context_source: str,
) -> str:
    zotero_abstract = zotero_item.abstract if zotero_item and zotero_item.abstract else ""
    zotero_title = zotero_item.title if zotero_item else ""
    zotero_doi = zotero_item.doi if zotero_item else ""
    zotero_date = zotero_item.date if zotero_item else ""
    return textwrap.dedent(
        f"""
        You are extracting metadata for a curated microplastics methods crosswalk.

        Fill three fields for the reference using ONLY the supplied
        Zotero/PDF/URL context. Do not use outside knowledge. If the evidence
        is absent or too ambiguous, use null for that field.

        Target fields:
        - scope: concise description of the scope of the study/method/document (e.g., "marine", "freshwater", "sediment", EU drinking water monitoring", "Intentionally added products", etc.)
        - particle_size_range: concise particle size range (report in μm) or operational size
          domain studied/reported. Include lower/upper bounds
          only if supported.
        - key_metrics_output: concise list of key quantitative outputs (e.g., mass, particle shape, size, density, polymer type, color, etc.),
          performance metrics, endpoints, or reported method outputs.
        - abstract: use null when Zotero abstractNote is populated. If Zotero
          abstractNote is absent and no formal abstract is visible in the PDF
          context, write a concise factual abstract-style summary without
          inventing unsupported results.

        Return strict JSON only with this schema. Do not wrap the JSON in
        Markdown code fences.
        {{
          "particle_size_range": string or null,
          "key_metrics_output": string or null,
          "abstract": string or null,
          "confidence": "high" | "medium" | "low",
          "evidence": {{
            "particle_size_range": [short supporting quote strings],
            "key_metrics_output": [short supporting quote strings],
            "abstract": [short supporting quote strings]
          }}
        }}

        Crosswalk row:
        Short Citation: {row.get("Short Citation") or ""}
        Year: {row.get("Year") or ""}
        Full Title: {row.get("Full Title") or ""}
        DOI/URL: {row.get("DOI/URL") or ""}
        Existing Particle Size Range: {row.get("Particle Size Range") or ""}
        Existing Key Metrics / Output: {row.get("Key Metrics / Output") or ""}

        Zotero metadata:
        Title: {zotero_title or ""}
        DOI: {zotero_doi or ""}
        Date: {zotero_date or ""}
        Zotero abstractNote, which will be used directly for the Abstract
        field when present:
        {zotero_abstract}

        Context source: {context_source}

        Source context:
        {context}
        """
    ).strip()


def anthropic_headers(api_key: str) -> dict[str, str]:
    return {
        "content-type": "application/json",
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
    }


def available_anthropic_models(api_key: str) -> list[str]:
    request = urllib.request.Request(
        "https://api.anthropic.com/v1/models",
        headers=anthropic_headers(api_key),
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        raw = json.loads(response.read().decode("utf-8"))
    return [model["id"] for model in raw.get("data", []) if model.get("id")]


def choose_fallback_model(model_ids: list[str]) -> str | None:
    for model_id in model_ids:
        if "haiku" in model_id.lower():
            return model_id
    return model_ids[0] if model_ids else None


def parse_json_response(text: str) -> dict[str, Any]:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.I)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            return json.loads(cleaned[start : end + 1])
        raise


def call_anthropic(prompt: str, model: str, max_tokens: int) -> dict[str, Any]:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY is not set.")

    return call_anthropic_once(prompt, model, max_tokens, api_key)


def call_anthropic_once(
    prompt: str,
    model: str,
    max_tokens: int,
    api_key: str,
    allow_model_fallback: bool = True,
) -> dict[str, Any]:

    payload = {
        "model": model,
        "max_tokens": max_tokens,
        "temperature": 0,
        "messages": [{"role": "user", "content": prompt}],
    }
    request = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=json.dumps(payload).encode("utf-8"),
        headers=anthropic_headers(api_key),
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            raw = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        if allow_model_fallback and exc.code == 404 and "model:" in detail:
            models = available_anthropic_models(api_key)
            fallback = choose_fallback_model(models)
            if fallback and fallback != model:
                print(f"  Model {model} unavailable; retrying with {fallback}")
                return call_anthropic_once(
                    prompt,
                    fallback,
                    max_tokens,
                    api_key,
                    allow_model_fallback=False,
                )
        raise RuntimeError(f"Anthropic API error {exc.code}: {detail}") from exc

    text = "".join(
        block.get("text", "")
        for block in raw.get("content", [])
        if block.get("type") == "text"
    ).strip()
    try:
        return parse_json_response(text)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Claude did not return valid JSON: {text[:1000]}") from exc


def row_dict(ws, row_idx: int, header_cols: dict[str, int]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for normalized, col in header_cols.items():
        header = ws.cell(HEADER_ROW, col).value
        if header:
            out[" ".join(str(header).replace("\n", " ").split())] = ws.cell(row_idx, col).value
    return out


def rows_to_process(ws, headers: dict[str, int], args: argparse.Namespace) -> list[int]:
    short_citation_col = headers.get(normalize_header("Short Citation"))
    doi_col = headers.get(normalize_header("DOI/URL"))
    title_col = headers.get(normalize_header("Full Title"))
    reviewed_col = headers.get(normalize_header("Reviewed"))
    exclude_rows = set(getattr(args, "exclude_rows", set()))
    start_row = max(DATA_START_ROW, args.start_row or DATA_START_ROW)
    candidate_rows = args.row if args.row else range(start_row, ws.max_row + 1)
    consecutive_empty_short_citations = 0
    rows: list[int] = []
    for row_idx in candidate_rows:
        if row_idx in exclude_rows:
            continue
        if short_citation_col:
            short_citation = str(ws.cell(row_idx, short_citation_col).value or "").strip()
            if not short_citation:
                consecutive_empty_short_citations += 1
                print(
                    f"Row {row_idx}: empty Short Citation; "
                    "skipping metadata extraction."
                )
                if consecutive_empty_short_citations >= 3:
                    print("Stopping after 3 consecutive empty Short Citation rows.")
                    break
                continue
            consecutive_empty_short_citations = 0
        if args.doi and doi_col:
            row_doi = normalize_doi(ws.cell(row_idx, doi_col).value)
            if row_doi != normalize_doi(args.doi):
                continue
        if args.title_contains and title_col:
            title = str(ws.cell(row_idx, title_col).value or "").lower()
            if args.title_contains.lower() not in title:
                continue
        if args.only_skipped and reviewed_col:
            reviewed = str(ws.cell(row_idx, reviewed_col).value or "").strip()
            if reviewed.lower() == "yes":
                continue
        rows.append(row_idx)
        if args.limit is not None and len(rows) >= args.limit:
            break
    return rows


def should_write(existing: Any, new_value: Any, overwrite: bool) -> bool:
    if new_value in (None, ""):
        return False
    if overwrite:
        return True
    return existing in (None, "")


def apply_zotero_abstract(result: dict[str, Any], item: ZoteroItem | None) -> dict[str, Any]:
    if item and item.abstract:
        result["abstract"] = item.abstract
        result.setdefault("evidence", {})
        result["evidence"]["abstract"] = ["Zotero abstractNote"]
    return result


def set_status_cells(
    ws,
    row_idx: int,
    target_cols: dict[str, int],
    pdf_available: str,
    reviewed: str,
    write: bool,
) -> None:
    if not write:
        return
    ws.cell(row_idx, target_cols["pdf_available_zotero"]).value = pdf_available
    ws.cell(row_idx, target_cols["reviewed"]).value = reviewed


def append_audit_rows(audit_csv: Path, audit_rows: list[dict[str, Any]]) -> None:
    audit_csv.parent.mkdir(parents=True, exist_ok=True)
    existing_rows: list[dict[str, Any]] = []
    existing_fieldnames: list[str] = []

    if audit_csv.exists() and audit_csv.stat().st_size > 0:
        with audit_csv.open("r", newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            existing_fieldnames = list(reader.fieldnames or [])
            existing_rows = list(reader)

    fieldnames: list[str] = []
    seen: set[str] = set()
    for key in existing_fieldnames:
        if key not in seen:
            fieldnames.append(key)
            seen.add(key)
    for row in audit_rows:
        for key in sorted(row):
            if key not in seen:
                fieldnames.append(key)
                seen.add(key)

    with audit_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(existing_rows)
        writer.writerows(audit_rows)


def apply_resume_defaults(args: argparse.Namespace, base_dir: Path) -> argparse.Namespace:
    if not args.resume:
        return args

    filled_workbook = base_dir / "crosswalk.metadata_filled.xlsx"
    if args.workbook == "crosswalk.xlsx" and filled_workbook.exists():
        args.workbook = str(filled_workbook)
    args.only_skipped = True
    args.write = True
    args.in_place = True
    return args


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", default="crosswalk.xlsx")
    parser.add_argument("--sheet", default="Crosswalk Table")
    parser.add_argument("--zotero-db", default=str(Path.home() / "Zotero" / "zotero.sqlite"))
    parser.add_argument("--zotero-storage", default=str(Path.home() / "Zotero" / "storage"))
    parser.add_argument("--collection-name", action="append", default=["Quality Standards for MNPs"])
    parser.add_argument("--collection-id", action="append", type=int, default=[])
    parser.add_argument("--doi")
    parser.add_argument("--title-contains")
    parser.add_argument("--row", action="append", type=int)
    parser.add_argument(
        "--start-row",
        type=int,
        help=(
            "Start scanning at this Excel row number. Combines with --resume, "
            "--batch-size, and --batch-loop. Ignored when --row is supplied."
        ),
    )
    parser.add_argument("--limit", type=int)
    parser.add_argument(
        "--resume",
        action="store_true",
        help=(
            "Resume from crosswalk.metadata_filled.xlsx if it exists, process "
            "only unreviewed rows, and write changes in place."
        ),
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        help="Alias for --limit; intended for resume/batch processing.",
    )
    parser.add_argument(
        "--batch-loop",
        action="store_true",
        help="Keep processing batches until no selected rows remain.",
    )
    parser.add_argument(
        "--only-skipped",
        action="store_true",
        help=(
            "Only process rows whose Reviewed column is not Yes. "
            "Useful after adding PDFs for previously skipped rows."
        ),
    )
    parser.add_argument("--cache-dir", default="metadata_cache")
    parser.add_argument("--audit-csv", default="metadata_cache/crosswalk_metadata_audit.csv")
    parser.add_argument("--env-file", default=".env")
    parser.add_argument("--model", default=os.environ.get("ANTHROPIC_MODEL", DEFAULT_MODEL))
    parser.add_argument("--max-input-chars", type=int, default=25000)
    parser.add_argument("--max-output-tokens", type=int, default=3000)
    parser.add_argument("--skip-llm", action="store_true")
    parser.add_argument(
        "--skip-url-scrape",
        action="store_true",
        help="Do not fetch DOI/URL landing pages when Zotero/PDF context is unavailable.",
    )
    parser.add_argument("--write", action="store_true", help="Write workbook output. Default is dry-run.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite non-empty target cells.")
    parser.add_argument("--output", help="Output workbook path. Defaults to a metadata_filled copy when --write is used.")
    parser.add_argument("--in-place", action="store_true", help="Write back to --workbook. Use with care.")
    return parser.parse_args()


def run_once(args: argparse.Namespace, base_dir: Path) -> list[int]:
    if args.batch_size is not None:
        args.limit = args.batch_size

    workbook_path = Path(args.workbook)
    if not workbook_path.is_absolute():
        workbook_path = base_dir / workbook_path
    cache_dir = Path(args.cache_dir)
    if not cache_dir.is_absolute():
        cache_dir = base_dir / cache_dir
    audit_csv = Path(args.audit_csv)
    if not audit_csv.is_absolute():
        audit_csv = base_dir / audit_csv

    con = sqlite3.connect(
        f"file:{Path(args.zotero_db).as_posix()}?mode=ro&immutable=1",
        uri=True,
    )
    cids = collection_ids(con, args.collection_name, args.collection_id)
    print(f"Zotero collections: {cids}")
    items = zotero_items(con, Path(args.zotero_storage), cids)
    print(f"Loaded {len(items)} Zotero items from selected collections.")

    wb = load_workbook(workbook_path)
    ws = wb[args.sheet]
    target_cols = ensure_target_columns(ws)
    headers = workbook_headers(ws)
    rows = rows_to_process(ws, headers, args)
    print(f"Rows selected: {rows}")

    audit_rows: list[dict[str, Any]] = []
    for row_idx in rows:
        row = row_dict(ws, row_idx, headers)
        item = choose_item(row, items)
        audit: dict[str, Any] = {
            "row": row_idx,
            "short_citation": row.get("Short Citation"),
            "doi_url": row.get("DOI/URL"),
            "source_url": source_url_from_row(row),
            "matched_item_id": item.item_id if item else None,
            "matched_item_key": item.key if item else None,
            "pdf": None,
            "url_cache": None,
            "status": None,
            "confidence": None,
        }
        markdown = None
        url_context = None
        pdf_available = "Yes" if item and item.attachment_paths else "No"

        if item is None:
            print(f"Row {row_idx}: no Zotero match")
            if not args.skip_url_scrape:
                try:
                    url_context, url_source, url_cache = scrape_landing_page_context(
                        audit["source_url"],
                        cache_dir / "url_context",
                        args.max_input_chars,
                    )
                    audit["url_cache"] = url_cache
                    audit["url_context_source"] = url_source
                except Exception as exc:
                    audit["url_context_source"] = f"url_scrape_failed: {exc}"
                    print(f"  URL scrape failed: {exc}")
            if not url_context:
                audit["status"] = "no_zotero_match"
                set_status_cells(
                    ws,
                    row_idx,
                    target_cols,
                    pdf_available="No",
                    reviewed="No - no Zotero match",
                    write=args.write,
                )
                audit_rows.append(audit)
                continue
            print(f"  Using DOI/URL landing-page context: {audit['source_url']}")

        elif not item.abstract and not item.attachment_paths:
            print(
                f"Row {row_idx}: matched Zotero item {item.key}, "
                "but no Zotero abstract or PDF"
            )
            if not args.skip_url_scrape:
                try:
                    url_context, url_source, url_cache = scrape_landing_page_context(
                        audit["source_url"],
                        cache_dir / "url_context",
                        args.max_input_chars,
                    )
                    audit["url_cache"] = url_cache
                    audit["url_context_source"] = url_source
                except Exception as exc:
                    audit["url_context_source"] = f"url_scrape_failed: {exc}"
                    print(f"  URL scrape failed: {exc}")
            if not url_context:
                audit["status"] = "no_pdf_attachment"
                set_status_cells(
                    ws,
                    row_idx,
                    target_cols,
                    pdf_available="No",
                    reviewed="No - no abstract/PDF/URL context",
                    write=args.write,
                )
                audit_rows.append(audit)
                continue
            print(f"  Using DOI/URL landing-page context: {audit['source_url']}")

        pdf_path = item.attachment_paths[0] if item and item.attachment_paths else None
        audit["pdf"] = str(pdf_path) if pdf_path else None
        if item and item.abstract:
            print(
                f"Row {row_idx}: matched {item.key}; "
                f"using Zotero abstractNote; PDF available={pdf_available}"
            )
        elif pdf_path:
            print(f"Row {row_idx}: matched {item.key}; PDF={pdf_path.name}")
            try:
                md_path = convert_pdf_to_markdown(pdf_path, cache_dir / "markdown")
                markdown = md_path.read_text(encoding="utf-8", errors="replace")
            except Exception as exc:
                audit["status"] = f"pdf_conversion_failed: {exc}"
                print(f"  PDF conversion failed: {exc}")
                set_status_cells(
                    ws,
                    row_idx,
                    target_cols,
                    pdf_available=pdf_available,
                    reviewed="No - PDF conversion failed",
                    write=args.write,
                )
                audit_rows.append(audit)
                continue

        context, context_source = resolve_llm_context(
            item, markdown, url_context, args.max_input_chars
        )
        audit["context_source"] = context_source

        if args.skip_llm:
            result = {
                "particle_size_range": None,
                "key_metrics_output": None,
                "abstract": item.abstract if item else None,
                "confidence": "low",
                "evidence": {},
            }
            audit["status"] = "skipped_llm"
            reviewed_status = "No - LLM skipped"
        else:
            prompt = build_prompt(
                row,
                item,
                context,
                context_source,
            )
            source_key = item.key if item else "url"
            prompt_path = cache_dir / "prompts" / f"row_{row_idx}_{source_key}.txt"
            prompt_path.parent.mkdir(parents=True, exist_ok=True)
            prompt_path.write_text(prompt, encoding="utf-8")
            try:
                result = call_anthropic(prompt, args.model, args.max_output_tokens)
                audit["status"] = "llm_ok"
                reviewed_status = "Yes"
            except Exception as exc:
                audit["status"] = f"llm_failed: {exc}"
                print(f"  LLM failed: {exc}")
                set_status_cells(
                    ws,
                    row_idx,
                    target_cols,
                    pdf_available=pdf_available,
                    reviewed="No - LLM failed",
                    write=args.write,
                )
                audit_rows.append(audit)
                continue

        result = apply_zotero_abstract(result, item)
        audit["confidence"] = result.get("confidence")
        for key in ["particle_size_range", "key_metrics_output", "abstract"]:
            col = target_cols[key]
            value = result.get(key)
            existing = ws.cell(row_idx, col).value
            audit[f"proposed_{key}"] = value
            audit[f"existing_{key}"] = existing
            if args.write and should_write(existing, value, args.overwrite):
                ws.cell(row_idx, col).value = value

        set_status_cells(
            ws,
            row_idx,
            target_cols,
            pdf_available=pdf_available,
            reviewed=reviewed_status,
            write=args.write,
        )
        audit["proposed_pdf_available_zotero"] = pdf_available
        audit["proposed_reviewed"] = reviewed_status

        source_key = item.key if item else "url"
        evidence_path = cache_dir / "responses" / f"row_{row_idx}_{source_key}.json"
        evidence_path.parent.mkdir(parents=True, exist_ok=True)
        evidence_path.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
        audit["response_json"] = str(evidence_path)
        audit_rows.append(audit)

        print(json.dumps(result, indent=2, ensure_ascii=False))

    append_audit_rows(audit_csv, audit_rows)
    print(f"Audit CSV: {audit_csv}")

    if args.write:
        if args.in_place:
            output_path = workbook_path
            backup_path = workbook_path.with_suffix(".before_metadata_fill.xlsx")
            if not backup_path.exists():
                shutil.copy2(workbook_path, backup_path)
                print(f"Backup written: {backup_path}")
        else:
            output_path = Path(args.output) if args.output else workbook_path.with_name(
                f"{workbook_path.stem}.metadata_filled{workbook_path.suffix}"
            )
            if not output_path.is_absolute():
                output_path = base_dir / output_path
        wb.save(output_path)
        print(f"Workbook written: {output_path}")
    else:
        print("Dry run only. Re-run with --write to create workbook output.")

    return rows


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    base_dir = Path(__file__).resolve().parent
    early_args = argparse.ArgumentParser(add_help=False)
    early_args.add_argument("--env-file", default=".env")
    env_args, _ = early_args.parse_known_args()
    env_path = Path(env_args.env_file)
    if not env_path.is_absolute():
        env_path = base_dir / env_path
    load_env_file(env_path)

    args = apply_resume_defaults(parse_args(), base_dir)
    if args.batch_size is not None:
        args.limit = args.batch_size

    attempted_rows: set[int] = set()
    while True:
        run_args = argparse.Namespace(**vars(args))
        run_args.exclude_rows = attempted_rows
        selected_rows = run_once(run_args, base_dir)
        attempted_rows.update(selected_rows)
        if not args.batch_loop or not selected_rows:
            break
        print(f"Completed batch of {len(selected_rows)}; starting next batch.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
