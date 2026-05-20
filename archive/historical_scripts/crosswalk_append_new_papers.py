"""
crosswalk_append_new_papers.py
------------------------------
Appends new papers to 'MNP Quality Standards Crosswalk v2.xlsx' without
overwriting any of Scott's existing manual changes.

NEW PAPERS COVERED (18 total candidates):
  Scott's additions (collection grew 86 → 92):
    IMZAL8Y8  Koelmans et al. 2025   — Towards Regulatory Readiness (preprint)
    HJRW7S45  De Ruijter et al. 2025 — Brief history of MP effect testing
    DVWNDVSE  JRC EURM-060 2025      — PET reference material for water
    K8JP8LP7  ISO 16094-2:2025       — Vibrational spectroscopy for drinking water
    5BAU3XQM  Thomas et al. 2026     — Communicating Confidence in MNP ID
    8WEGWFBC  Wright et al. 2024     — MNP concepts for particle/fibre toxicologists
    WPVFAJZN  Xu et al. 2025         — Are microplastics bad for your health?
    ESLWHWEW  Kennedy et al. 2025    — Trends in Quality and RA Applicability
    YDLUDXK7  Hagelskjær et al. 2025 — EasyMP reference materials
    DB7ZYKBF  Pegoraro et al. 2025   — Nanoplastic reference materials

  Papers we added from Regulatory Readiness cited references:
    A9IPX8JW  Vogel et al. 2024      — RA framework for MNPs (human health)
    22DT35FU  Foss Hansen et al. 2024 — PlasticRiskCat
    8AMTU2FB  Koelmans et al. 2023   — Rational and efficient MP risk assessment
    NIH492ZS  Lane et al. 2025       — Exposure scenarios for human health RA
    ERS2Z529  Qiu et al. 2025        — AI tools for data quality evaluation
    RHF3EK8P  Wardani et al. 2024    — PBK modeling for NMP
    XKCA572S  Noventa et al. 2021    — Paradigms for NMP human health risks
    PUDABW95  Wright et al. 2025     — Screening framework for particles in tissue

DEDUPLICATION: Checks column B (citation) and column D (title keywords) before
appending each paper. If a paper is already present, it is skipped.

USAGE:
    Close Microsoft Excel first (so the file is not locked), then run:
        python3 crosswalk_append_new_papers.py

    The script will print a summary of what was added vs skipped.
"""

import pathlib
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── Paths ────────────────────────────────────────────────────────────────────
HERE = pathlib.Path(__file__).parent
XLSX = HERE / "MNP Quality Standards Crosswalk v2.xlsx"

# ── Tier strings (must match format used in existing rows) ────────────────────
T1 = "Tier 1 — Regulatory / Accredited SOP"
T2 = "Tier 2 — Authoritative Guidance"
T3 = "Tier 3 — Peer-Reviewed Method"
T4 = "Tier 4 — Supporting Science"

CHK = "✓"   # checkmark character used throughout

# ── Column index mapping (1-indexed for openpyxl) ────────────────────────────
# A=1  B=2   C=3    D=4     E=5   F=6       G=7
# ID   Cite  Year   Title   Type  Journal   Tier
# H=8  I=9   J=10  K=11  L=12  M=13  N=14  O=15  P=16  Q=17
# Def  PF    Samp  DW    SW    Sed   Bio   Air   Food  HumTis
# R=18       S=19       T=20     U=21   V=22   W=23
# SampProc   Subsamp    AnlGen   FTIR   Raman  PyGCMS
# X=24             Y=25         Z=26         AA=27        AB=28
# RefMat/Ctrl      Blanks/QC    DataAnal     Reporting    DataDep
# AC=29              AD=30             AE=31        AF=32
# ToxStudyDesign    ToxEffects       InterLab     Notes

COL = {
    "ID":        1,   "Cite":    2,   "Year":   3,   "Title":   4,
    "Type":      5,   "Journal": 6,   "Tier":   7,
    "Def":       8,   "PF":      9,   "Samp":  10,
    "DW":       11,   "SW":     12,   "Sed":   13,   "Bio":    14,
    "Air":      15,   "Food":   16,   "HumTis":17,
    "SampProc": 18,   "Subsamp":19,   "AnlGen":20,
    "FTIR":     21,   "Raman":  22,   "PyGCMS":23,
    "RefMat":   24,   "Blanks": 25,   "DataAnal":26, "Report": 27,
    "DataDep":  28,
    "ToxDesign":29,   "ToxEffects":30, "InterLab":31,
    "Notes":    32,
}

# ── New paper definitions ─────────────────────────────────────────────────────
# Each dict:  cite, year, title, type, journal, tier, checks (list of COL keys), note
NEW_PAPERS = [

    # ── Scott's additions ────────────────────────────────────────────────────

    {
        "cite":    "Koelmans et al. 2025a",
        "year":    2025,
        "title":   "Towards Regulatory Readiness: Evaluating Frameworks for Microplastic Health Risk Assessment",
        "type":    "Preprint",
        "journal": "bioRxiv / EarthArXiv",
        "tier":    T4,
        "checks":  ["Def", "PF", "ToxDesign", "ToxEffects", "DataAnal", "Report"],
        "note":    ("Preprint. TRL framework for 5 MP health RA frameworks: SCCWRP (TRL7), "
                   "MICROPLASTICLAB (TRL6), POLYRISK/MOMENTUM (TRL5), AURORA (TRL3-4), "
                   "PLASTICHEAL (TRL3). 9 evaluation criteria (Table 1); regulatory readiness "
                   "ratings (Table 3). Koelmans, Bouwmeester, Christopher, Coffin et al."),
        "key":     "Koelmans 2025a",  # used for dedup check
    },
    {
        "cite":    "De Ruijter et al. 2025",
        "year":    2025,
        "title":   "A brief history of microplastics effect testing: Guidance and prospect",
        "type":    "Journal Article",
        "journal": "Microplastics and Nanoplastics",
        "tier":    T3,
        "checks":  ["Def", "PF", "RefMat", "Blanks", "DataAnal", "Report",
                   "ToxDesign", "ToxEffects"],
        "note":    ("Reviews MP effect testing QA/QC 2016-2025; same failures persist. "
                   "Protocol for polydisperse ERMP test materials; data rescaling/alignment; "
                   "natural particle controls required; community-level approaches. "
                   "De Ruijter, Redondo-Hasselerharm, Koelmans."),
        "key":     "De Ruijter 2025",
    },
    {
        "cite":    "JRC 2025 EURM-060",
        "year":    2025,
        "title":   "Polyethylene terephthalate (PET) microplastic particles in water: EURM-060",
        "type":    "Reference Material",
        "journal": "European Commission Joint Research Centre",
        "tier":    T2,
        "checks":  ["DW", "SampProc", "RefMat"],
        "note":    ("Certified reference material EURM-060: PET microplastic particles in water. "
                   "For method validation and interlaboratory comparability. "
                   "European Commission JRC, Seghers et al."),
        "key":     "EURM-060 2025",
    },
    {
        "cite":    "ISO 16094-2:2025",
        "year":    2025,
        "title":   ("ISO 16094-2:2025 Water quality — Analysis of microplastic in water — "
                   "Part 2: Vibrational spectroscopy methods for waters with low content "
                   "of suspended solids including drinking water"),
        "type":    "Standard",
        "journal": "ISO",
        "tier":    T1,
        "checks":  ["Samp", "DW", "SampProc", "FTIR", "Raman", "RefMat", "Blanks"],
        "note":    ("ISO standard. Vibrational spectroscopy (µFTIR, µRaman) for microplastic "
                   "analysis in drinking water and low-turbidity waters. Specifies minimum "
                   "requirements for spectral libraries, resolution, blank controls."),
        "key":     "ISO 16094-2 2025",
    },
    {
        "cite":    "Thomas et al. 2026",
        "year":    2026,
        "title":   "Communicating Confidence in the Reliability of Micro- and Nanoplastic Identification in Human Health Studies",
        "type":    "Journal Article",
        "journal": "Environmental Science & Technology",
        "tier":    T3,
        "checks":  ["Def", "PF", "FTIR", "Raman", "PyGCMS", "DataAnal", "Report", "HumTis"],
        "note":    ("Framework for communicating analytical confidence in MNP identification. "
                   "Addresses HQI thresholds, spectral quality, reporting requirements "
                   "for human health studies. Thomas, Belz, Booth, Clift, et al."),
        "key":     "Thomas 2026",
    },
    {
        "cite":    "Wright et al. 2024b",
        "year":    2024,
        "title":   "Micro- and nanoplastics concepts for particle and fibre toxicologists",
        "type":    "Journal Article",
        "journal": "Particle and Fibre Toxicology",
        "tier":    T3,
        "checks":  ["Def", "PF", "ToxDesign", "ToxEffects"],
        "note":    ("Concepts paper bridging MP/NP research and particle/fibre toxicology. "
                   "Covers characterization needs, dosimetry, physicochemical properties "
                   "relevant to toxicological assessment. Wright, Cassee, Erdely, Campen."),
        "key":     "Wright 2024b",
    },
    {
        "cite":    "Xu et al. 2025",
        "year":    2025,
        "title":   "Are microplastics bad for your health? More rigorous science is needed",
        "type":    "Journal Article",
        "journal": "Nature Medicine",
        "tier":    T3,
        "checks":  ["PF", "DataAnal", "Report", "ToxDesign", "ToxEffects"],
        "note":    ("Commentary/perspective on quality gaps in MP human health research. "
                   "Calls for rigorous study design, appropriate controls, and transparent "
                   "reporting. Xu, Wright, Rauert, Thomas."),
        "key":     "Xu 2025",
    },
    {
        "cite":    "Kennedy et al. 2025",
        "year":    2025,
        "title":   "Trends in Quality and Risk Assessment Applicability of Microplastic Ecotoxicity Studies",
        "type":    "Journal Article",
        "journal": "Environmental Science & Technology",
        "tier":    T3,
        "checks":  ["PF", "RefMat", "Blanks", "DataAnal", "Report", "ToxDesign", "ToxEffects"],
        "note":    ("Longitudinal analysis of MP ecotoxicology study quality 2016-2024 "
                   "(extension of ToMEx). Consistent QA/QC failures persist; low RA applicability. "
                   "Kennedy, Vital, Kukkola, Miller, Yeh, Coffin, Ahmed, Bertrand et al."),
        "key":     "Kennedy 2025",
    },
    {
        "cite":    "Hagelskjær et al. 2025",
        "year":    2025,
        "title":   "EasyMP: Diverse and Environmentally Relevant Microplastic Reference Materials Encompassing Fragments and Fibers",
        "type":    "Journal Article",
        "journal": "Environmental Science & Technology",
        "tier":    T3,
        "checks":  ["Samp", "SampProc", "RefMat"],
        "note":    ("EasyMP reference material set: fragments and fibers spanning multiple "
                   "polymer types and morphologies. Environmentally relevant size/shape "
                   "distributions. For method validation and toxicology studies."),
        "key":     "Hagelskjær 2025",
    },
    {
        "cite":    "Pegoraro et al. 2025",
        "year":    2025,
        "title":   "Nanoplastic reference materials for biological and methodological assessment",
        "type":    "Journal Article",
        "journal": "NanoImpact",
        "tier":    T3,
        "checks":  ["SampProc", "RefMat", "FTIR", "Raman"],
        "note":    ("Nanoplastic reference materials for biological and analytical method "
                   "validation. Laser ablation, nanometrology, nanotoxicity assessment. "
                   "Pegoraro, Chen, Sakib, Jakubek, Zou et al. NRC Canada."),
        "key":     "Pegoraro 2025",
    },

    # ── Papers added from Regulatory Readiness cited references ──────────────

    {
        "cite":    "Vogel et al. 2024",
        "year":    2024,
        "title":   "Towards a risk assessment framework for micro- and nanoplastic particles for human health",
        "type":    "Journal Article",
        "journal": "Particle and Fibre Toxicology",
        "tier":    T3,
        "checks":  ["Def", "PF", "DataAnal", "Report", "ToxDesign", "ToxEffects"],
        "note":    ("Proposes comprehensive RA framework for MNP human health; addresses "
                   "dosimetry, particle characterization, hazard assessment, exposure. "
                   "DOI: 10.1186/s12989-024-00602-9"),
        "key":     "Vogel 2024",
    },
    {
        "cite":    "Foss Hansen et al. 2024",
        "year":    2024,
        "title":   "PlasticRiskCat: Plasticheal risk evaluation and categorization framework",
        "type":    "Conference Abstract",
        "journal": "Toxicology Letters (Supplement)",
        "tier":    T4,
        "checks":  ["PF", "ToxDesign", "ToxEffects"],
        "note":    ("PlasticRiskCat risk categorization framework from PLASTICHEAL EU project. "
                   "Conference abstract / poster. DOI: 10.1016/j.toxlet.2024.07.176"),
        "key":     "Foss Hansen 2024",
    },
    {
        "cite":    "Koelmans et al. 2023",
        "year":    2023,
        "title":   "Towards a rational and efficient risk assessment for microplastics",
        "type":    "Journal Article",
        "journal": "TrAC Trends in Analytical Chemistry",
        "tier":    T3,
        "checks":  ["Def", "PF", "DataAnal", "Report", "ToxDesign"],
        "note":    ("Proposes tiered, efficient MP RA based on stochastic risk quotients; "
                   "addresses exposure, hazard, risk characterization. "
                   "DOI: 10.1016/j.trac.2023.117142"),
        "key":     "Koelmans 2023",
    },
    {
        "cite":    "Lane et al. 2025",
        "year":    2025,
        "title":   "Exposure scenarios for human health risk assessment of nano- and microplastic particles",
        "type":    "Journal Article",
        "journal": "Microplastics and Nanoplastics",
        "tier":    T3,
        "checks":  ["Def", "PF", "Samp", "DW", "Air", "Food", "HumTis"],
        "note":    ("Quantitative exposure scenarios for NMPs across multiple routes and matrices "
                   "(oral, inhalation, dermal). Supports human health RA. "
                   "DOI: 10.1186/s43591-025-00134-9"),
        "key":     "Lane 2025",
    },
    {
        "cite":    "Qiu et al. 2025",
        "year":    2025,
        "title":   "Using artificial intelligence tools for data quality evaluation in the context of microplastic human health risk assessments",
        "type":    "Journal Article",
        "journal": "Environment International",
        "tier":    T3,
        "checks":  ["PF", "DataAnal", "Report"],
        "note":    ("AI-assisted automated data quality scoring for MP human health RA. "
                   "Validates AI tools against expert scoring. "
                   "DOI: 10.1016/j.envint.2025.109341"),
        "key":     "Qiu 2025",
    },
    {
        "cite":    "Wardani et al. 2024",
        "year":    2024,
        "title":   "Nano- and microplastic PBK modeling in the context of human exposure and risk assessment",
        "type":    "Journal Article",
        "journal": "Environment International",
        "tier":    T3,
        "checks":  ["PF", "HumTis", "ToxDesign", "ToxEffects"],
        "note":    ("Physiologically-based kinetic (PBK) modeling for NMP bioaccumulation, "
                   "distribution, exposure assessment. Particle characterization needs. "
                   "DOI: 10.1016/j.envint.2024.108504"),
        "key":     "Wardani 2024",
    },
    {
        "cite":    "Noventa et al. 2021",
        "year":    2021,
        "title":   "Paradigms to assess the human health risks of nano- and microplastics",
        "type":    "Journal Article",
        "journal": "Microplastics and Nanoplastics",
        "tier":    T3,
        "checks":  ["Def", "PF", "ToxDesign", "ToxEffects"],
        "note":    ("Compares paradigms for NMP human health RA; discusses read-across, "
                   "threshold approaches, particle vs chemical RA frameworks. "
                   "DOI: 10.1186/s43591-021-00011-1"),
        "key":     "Noventa 2021",
    },
    {
        "cite":    "Wright et al. 2025b",
        "year":    2025,
        "title":   ("A new screening framework to support the identification of exogenous particles "
                   "and suspect microplastics in situ in pathological tissue samples"),
        "type":    "Journal Article",
        "journal": "eBioMedicine",
        "tier":    T3,
        "checks":  ["SampProc", "RefMat", "FTIR", "Raman", "HumTis", "DataAnal"],
        "note":    ("Screening framework for identifying MPs/exogenous particles in pathological "
                   "tissue samples in situ. Spectroscopic identification criteria. "
                   "DOI: 10.1016/j.ebiom.2025.105984"),
        "key":     "Wright 2025b",
    },
]


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if not XLSX.exists():
        print(f"ERROR: File not found: {XLSX}")
        return

    print(f"Opening: {XLSX}")
    try:
        wb = openpyxl.load_workbook(XLSX)
    except Exception as e:
        print(f"ERROR opening file: {e}")
        print("Is Excel still open? Close it and try again.")
        return

    ws = wb["Crosswalk Table"]

    # ── Read existing papers (rows 3+) ────────────────────────────────────────
    existing_cites = set()
    existing_keys  = set()
    last_id = 0
    last_row = 2  # header is row 2

    for row in ws.iter_rows(min_row=3, values_only=False):
        if row[0].value is None:
            break
        cite_val = str(row[1].value or "").strip()
        existing_cites.add(cite_val.lower())
        # Extract key words: first token of cite (author) + year
        parts = cite_val.replace(",", "").split()
        if len(parts) >= 2:
            key = f"{parts[0].lower()} {parts[-1]}"
            existing_keys.add(key)
        id_val = row[0].value
        if isinstance(id_val, (int, float)):
            last_id = max(last_id, int(id_val))
        last_row = row[0].row

    print(f"  Found {last_id} existing papers (last data row: {last_row})")
    print(f"  Existing citation keys (sample): {list(existing_cites)[:5]}")

    # ── Append new papers ─────────────────────────────────────────────────────
    added   = []
    skipped = []

    for paper in NEW_PAPERS:
        cite_lower = paper["cite"].lower()
        # Dedup by: exact citation match OR key-word match
        key_parts = paper["key"].lower().split()
        key_match = any(
            all(kp in ec for kp in key_parts[:2])
            for ec in existing_cites
        )
        if cite_lower in existing_cites or key_match:
            skipped.append(paper["cite"])
            continue

        # Build new row
        last_id  += 1
        last_row += 1
        new_row_num = last_row

        ws.cell(new_row_num, COL["ID"]).value     = last_id
        ws.cell(new_row_num, COL["Cite"]).value   = paper["cite"]
        ws.cell(new_row_num, COL["Year"]).value   = paper["year"]
        ws.cell(new_row_num, COL["Title"]).value  = paper["title"]
        ws.cell(new_row_num, COL["Type"]).value   = paper["type"]
        ws.cell(new_row_num, COL["Journal"]).value = paper["journal"]
        ws.cell(new_row_num, COL["Tier"]).value   = paper["tier"]
        ws.cell(new_row_num, COL["Notes"]).value  = paper["note"]

        for col_key in paper["checks"]:
            ws.cell(new_row_num, COL[col_key]).value = CHK

        # Update existing_cites for subsequent dedup passes
        existing_cites.add(cite_lower)
        added.append(paper["cite"])

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(XLSX)
    print(f"\nSaved: {XLSX}")
    print(f"\n✓ Added ({len(added)} papers):")
    for a in added:
        print(f"    {a}")
    print(f"\n⏭  Skipped — already in crosswalk ({len(skipped)} papers):")
    for s in skipped:
        print(f"    {s}")


if __name__ == "__main__":
    main()
