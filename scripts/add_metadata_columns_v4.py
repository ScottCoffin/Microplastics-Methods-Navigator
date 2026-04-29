"""
add_metadata_columns_v4.py
---------------------------
Creates MNP Quality Standards Crosswalk v4.xlsx by adding two new metadata columns
to the existing v3 crosswalk:

  Column AH (index 33): Particle Size Range
    — particle size range covered by the analytical method(s) in this paper

  Column AI (index 34): Key Metrics / Output
    — primary measurement outputs (particle count, mass, polymer type, etc.)

Data populated from PAPER_METADATA dict (keyed by crosswalk row ID), derived from:
  - Zotero full-text retrieval (2026-04-23)
  - Key Notes column
  - Domain knowledge

Usage:
    python add_metadata_columns_v4.py
Output:
    MNP Quality Standards Crosswalk v4.xlsx  (working folder)
"""

import pathlib, shutil
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

HERE = pathlib.Path(__file__).parent
SRC  = HERE / "MNP Quality Standards Crosswalk v3.xlsx"
DST  = HERE / "MNP Quality Standards Crosswalk v4.xlsx"

# ── Paper metadata keyed by crosswalk row ID ──────────────────────────────────
# size_range: particle size range covered by the paper's analytical method(s)
# key_metrics: primary measurement outputs
PAPER_METADATA = {
    # --- Drinking Water ---
    82: {
        "size_range": "\u22655 \u00b5m (FTIR / SWB-MP1); \u22651 \u00b5m (Raman / SWB-MP2)",
        "key_metrics": "Particle count/L by polymer type, size bin, morphology; blank-corrected",
    },
    85: {
        "size_range": "\u22655 \u00b5m (\u00b5FTIR; California DW extraction SOP)",
        "key_metrics": "Count/L by polymer/size/morphology; blank-corrected; LOD = mean(blanks) + 3 SD",
    },
    89: {
        "size_range": "\u226520 \u00b5m (\u00b5FTIR); \u22651 \u00b5m (\u00b5Raman) for DW and low-turbidity waters",
        "key_metrics": "Count/L, polymer type (10 priority polymers), size distribution, spectral match quality (HQI)",
    },
    102: {
        "size_range": "20\u20135000 \u00b5m (particles, area-equivalent diam.); 20\u201315000 \u00b5m (fibres, length); 5 size bins",
        "key_metrics": "Count/m\u00b3 by polymer type; \u22651000 L sample; 10 procedural blanks; spike recovery 100\u00b140%",
    },
    23: {
        "size_range": "1\u201320, 20\u2013212, 212\u2013500, >500 \u00b5m (four fractions spiked; 22-lab ILC)",
        "key_metrics": "Recovery 76\u00b110% SE (\u226520 \u00b5m: 92\u00b112%); FTIR accuracy 95%, Raman 91%; blank 91\u00b1141 particles",
    },
    77: {
        "size_range": "Not specified (risk assessment framework; no primary analytical SOP)",
        "key_metrics": "Qualitative risk characterization; exposure and hazard assessment framework",
    },
    56: {
        "size_range": ">20 \u00b5m (density separation validation in DW matrix)",
        "key_metrics": "Extraction efficiency (spike-recovery); EURM-060 reference material validation",
    },
    88: {
        "size_range": "~10 \u00b5m PET particles in water matrix (certified reference material EURM-060)",
        "key_metrics": "Certified count-based reference material for extraction/spectroscopy method validation",
    },
    68: {
        "size_range": "Multiple size fractions tested across DW and SW matrices",
        "key_metrics": "Extraction efficiency by matrix; method comparison; background particle levels",
    },
    16: {
        "size_range": "\u22655 \u00b5m (DW monitoring context; CA SOP-based)",
        "key_metrics": "Risk characterization; integrated monitoring + RA framework; used in CA DW regulation",
    },
    # --- Surface Water / Wastewater ---
    64: {
        "size_range": ">0.3 mm (manta trawl); <0.3 mm (pump sampling)",
        "key_metrics": "Count and polymer type in water column; multi-matrix comparison across SF Bay",
    },
    46: {
        "size_range": ">10 \u00b5m (tire wear particles focus; estimated practical lower limit)",
        "key_metrics": "Tire wear particle count and polymer ID; morphology-specific criteria",
    },
    47: {
        "size_range": "Not specified (risk framework based on environmental monitoring data)",
        "key_metrics": "Action levels for aquatic MPs; risk quotient framework; monitoring data integration",
    },
    1: {
        "size_range": "Any (bulk pyrolysis; no particle size threshold \u2014 reports total polymer mass)",
        "key_metrics": "Mass/L by polymer type (\u00b5g/L); 92\u201399% removal in drinking water treatment demonstrated",
    },
    62: {
        "size_range": "\u22651 \u00b5m (Raman); \u226510\u201320 \u00b5m (FTIR practical limit) for clean/surface water",
        "key_metrics": "Count/volume by polymer/size; spectral HQI >70%; library \u226510 polymer types, \u22657 size bins",
    },
    # --- Sediment ---
    76: {
        "size_range": "Various (sediment transport theory applied to MPs; size-dependent settling velocity)",
        "key_metrics": "Settling velocity; density-based fractionation; methodological context for MP behavior in sediment",
    },
    41: {
        "size_range": "Spiked MPs (µm range, not precisely constrained); marine/estuarine sediment matrix",
        "key_metrics": "Extraction recovery: core 47%, augmented 78%; FTIR ID accuracy 99.8%; Raman 96.7%; 2-lab ILC",
    },
    # --- Sediment + Biota ---
    103: {
        "size_range": "\u22651 \u00b5m (operationally defined by filter pore size after density separation/digestion)",
        "key_metrics": "Count/kg dw (sediment); count/g ww (biota tissue); MDA blank-correction (Lao 2025); Tier 1 CA field SOP",
    },
    104: {
        "size_range": "\u2265330 \u00b5m (manta net, surface water); \u22651.2 \u00b5m (GF/C filter, sediment & fish GIT)",
        "key_metrics": "Count/m\u00b3 (water), /g dw (sediment), /fish; 12-polymer library; \u00b5Raman 785 nm; recovery \u226580%",
    },
    # --- Biota ---
    58: {
        "size_range": "\u22650.3 mm (storm-petrels); \u22651 mm (most species); \u22655 mm (mesoplastics); sieve stack 0.3/1/5 mm",
        "key_metrics": "Frequency of occurrence; count and mass/individual; GIT compartment-specific; size class and shape",
    },
    69: {
        "size_range": "PE/PP/PET spike particles (\u00b5m range); fish GIT and mussel tissue matrices",
        "key_metrics": "KOH digestion recovery 96.7%; H\u2082O\u2082 recovery 88.8%; interlaboratory CV <11%; 3 polymers; 4 labs",
    },
    24: {
        "size_range": "Not specified (QA framework for biota biomonitoring; method-agnostic)",
        "key_metrics": "Quality score rubric (0\u2013100 scale); 18 criteria for biota biomonitoring study quality",
    },
    73: {
        "size_range": "Not specified (TRL assessment; technology readiness evaluation, not primary methods paper)",
        "key_metrics": "Technology readiness levels (TRL 1\u20139) for biota monitoring methods",
    },
    # --- Air / Atmospheric ---
    79: {
        "size_range": ">10 \u00b5m (majority of 27 reviewed studies); recommends prioritizing <10 \u00b5m for human health",
        "key_metrics": "Count/m\u00b3 or deposition rate (particles/m\u00b2/day); 11-criterion QA/QC score; mean 48.6% across studies",
    },
    100: {
        "size_range": "\u226510 \u00b5m (FPA-\u00b5FTIR; wet deposition filter samples)",
        "key_metrics": "Count and polymer type; ~90% spike-recovery; deposition flux (particles/m\u00b2/event or /day)",
    },
    101: {
        "size_range": "Various (review of passive and active air sampling studies; no single size standard)",
        "key_metrics": "Review of count, flux, mass metrics across atmospheric MP studies; standardization gaps identified",
    },
    # --- Food / Diet ---
    29: {
        "size_range": "0.001\u20135000 \u00b5m (NMP size range; dietary + inhalation multi-route framework)",
        "key_metrics": "Daily intake (particles/day, \u00b5g/day) by exposure route; hazard quotient for human health RA",
    },
    94: {
        "size_range": "0.1\u20135000 \u00b5m (NMPs; multi-route exposure scenarios \u2014 oral, inhalation, dermal)",
        "key_metrics": "Estimated daily intake by route and matrix; population exposure distributions",
    },
    # --- Human Tissue / Biomonitoring ---
    60: {
        "size_range": "Any (bulk Py-GC-MS; no particle size threshold \u2014 reports total polymer mass in blood)",
        "key_metrics": "Mass concentration (\u00b5g/mL) by polymer type in human blood; full method validation provided",
    },
    98: {
        "size_range": "\u22655 \u00b5m (\u00b5FTIR/\u00b5Raman for in situ tissue identification; pathological sections)",
        "key_metrics": "In situ particle identification and polymer confirmation in human pathological tissue",
    },
    96: {
        "size_range": "0.1\u201310 \u00b5m (NMPs; nano-range bioaccumulation modeling focus)",
        "key_metrics": "Tissue concentration estimates; bioaccumulation factors; organ clearance rates (PBK model)",
    },
}

# ── Load workbook ─────────────────────────────────────────────────────────────
shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)
ws = wb["Crosswalk Table"]

# ── Style helpers ─────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="D0E4FF")   # light blue for new cols
BOLD = Font(bold=True, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NEW_COLS = {
    34: ("Particle\nSize Range", "Particle size range covered by method(s) in this paper"),
    35: ("Key Metrics /\nOutput", "Primary measurement outputs (count, mass, polymer type, etc.)"),
}

# ── Add section header (row 1) ────────────────────────────────────────────────
ws.cell(row=1, column=34).value = "Extended Metadata"
ws.cell(row=1, column=34).font = Font(bold=True, size=10)
ws.cell(row=1, column=34).fill = HEADER_FILL

# ── Add column headers (row 2) ────────────────────────────────────────────────
for col_idx, (header, _) in NEW_COLS.items():
    c = ws.cell(row=2, column=col_idx)
    c.value = header
    c.font = BOLD
    c.fill = HEADER_FILL
    c.alignment = WRAP
    c.border = BORDER

# ── Set column widths ─────────────────────────────────────────────────────────
ws.column_dimensions["AH"].width = 38
ws.column_dimensions["AI"].width = 52

# ── Populate data rows ────────────────────────────────────────────────────────
filled = 0
skipped = 0
for row_obj in ws.iter_rows(min_row=3):
    pid_cell = row_obj[0]   # column A = row ID
    if pid_cell.value is None:
        continue
    try:
        pid = int(pid_cell.value)
    except (TypeError, ValueError):
        continue

    meta = PAPER_METADATA.get(pid)
    row_num = pid_cell.row

    size_cell   = ws.cell(row=row_num, column=34)
    metric_cell = ws.cell(row=row_num, column=35)

    if meta:
        size_cell.value   = meta.get("size_range", "")
        metric_cell.value = meta.get("key_metrics", "")
        filled += 1
    else:
        size_cell.value   = ""
        metric_cell.value = ""
        skipped += 1

    for cell in (size_cell, metric_cell):
        cell.alignment = WRAP
        cell.border = BORDER
        cell.font = Font(size=9)

# ── Freeze panes and save ─────────────────────────────────────────────────────
ws.freeze_panes = "C3"
wb.save(DST)

print(f"Saved: {DST}")
print(f"  Rows populated: {filled}")
print(f"  Rows without metadata: {skipped}")
print(f"  New columns: AH (Particle Size Range), AI (Key Metrics / Output)")
